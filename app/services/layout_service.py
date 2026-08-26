"""
Módulo de Layout — las ubicaciones se crean y gestionan 100% en el WMS.

Independiente de Reposición: este servicio es el cimiento (crear, clasificar y
poblar ubicaciones); Reposición, Picking, Recepción, Averías y Compras son
consumidores de ese cimiento, no dueños de la lógica de ubicaciones.

Dirección física de 5 ejes: Pasillo -> Fila (1/2, lado del pasillo) -> Cuerpo
(bahía) -> Nivel (entrepaño, 1 = piso) -> Hueco (espacio dentro del entrepaño).

Piezas:
  1. letras_disponibles()   — pasillos: A-Z y luego AA, AB... (revelado progresivo)
  2. crear_cuerpo()         — Mecanismo A: entrepaños+huecos de un Cuerpo, en bloque,
                              con Zona sugerida por Nivel (piso->PICKING, alto->RESERVA)
  3. crear_ubicacion_averias() — AVE1, AVE2... numeradas, fuera de la grilla pasillo/fila
  4. asignar_producto()     — Mecanismo B: bind ubicación↔SKU, con la regla 1:1 en PICKING/IMPORTADOS
  5. reclasificar_ubicacion() — cambiar zona/capacidad/activo/slot, con guardarraíles
  6. importar_excel()       — Mecanismo A, opción masiva (reusa asignar_producto)

editar_fila()/eliminar_fila() operan sobre el campo legado 'estante' (fila
plana previa a este rediseño) — siguen intactas para gestionar en bloque las
ubicaciones creadas antes del esquema de 5 ejes.

Operaciones a nivel de Cuerpo completo (todos sus Entrepaños/Huecos a la vez,
esquema de 5 ejes — no confundir con editar_fila()/eliminar_fila() del legado):
  - editar_cuerpo()        — remodula cantidad de entrepaños/huecos: borra y
                              reconstruye limpio, devolviendo stock a SIESA-GENERAL
                              antes de borrar. Bloquea si hay historial real.
  - eliminar_cuerpo()       — borra el cuerpo entero, todo o nada. Bloquea si
                              CUALQUIER hueco tiene stock activo o historial real.
  - reclasificar_cuerpo()   — cambia zona de todo el cuerpo, o lo desactiva
                              completo (camino recomendado para retirar un
                              cuerpo con historial sin perder la auditoría).

Todas las ubicaciones que crea este módulo nacen con origen='MANUAL' — el sync
de Siesa (ubicaciones_sync_service.py) nunca las toca.
"""
import re
from datetime import datetime
from app.extensions import db
from app.models.almacen import Almacen
from app.models.ubicacion import Ubicacion
from app.models.producto import Producto
from app.models.inventario import UbicacionProducto, MovimientoInventario
from app.models.picking import TareaPicking
from app.models.tarea_reposicion import TareaReposicion
from app.models.conteo import SesionConteo
from app.models.recepcion import ItemRecepcion
from app.models.lpn import LPN
from app.models.devolucion import TareaDevolucion
from app.models.devolucion_cliente import LineaDevolucionCliente


# ──────────────────────────────────────────────────────────────────────────────
# 1. Pasillos — A-Z y luego AA, AB... (esquema de columnas de Excel)
# ──────────────────────────────────────────────────────────────────────────────

def _generar_letras(cantidad: int):
    """A, B, ... Z, AA, AB, ... — bijective base-26, orden de generación real."""
    letras = []
    n = 1
    while len(letras) < cantidad:
        letras.append(_indice_a_letra(n))
        n += 1
    return letras


def _indice_a_letra(n: int) -> str:
    """1→A, 26→Z, 27→AA, 28→AB... (bijective base-26, sin dígito cero)."""
    letras = ''
    while n > 0:
        n, resto = divmod(n - 1, 26)
        letras = chr(65 + resto) + letras
    return letras


def letras_disponibles(almacen_id: int, minimo_a_mostrar: int = 5):
    """
    Devuelve las letras de pasillo aún no usadas en este almacén.
    Revelado progresivo: solo abre AA, AB... cuando las simples (A-Z) ya se agotaron.
    """
    usadas = {
        p for (p,) in db.session.query(Ubicacion.pasillo)
        .filter(Ubicacion.almacen_id == almacen_id, Ubicacion.pasillo.isnot(None))
        .distinct().all()
    }

    disponibles = []
    n = 1
    # Tope defensivo (26 simples + 676 dobles = 702) — nunca triples.
    while len(disponibles) < minimo_a_mostrar and n <= 702:
        letra = _indice_a_letra(n)
        if letra not in usadas:
            disponibles.append(letra)
        n += 1
    return disponibles


# ──────────────────────────────────────────────────────────────────────────────
# 2. Nivel 1 del Mecanismo A — crear las posiciones de una fila, en bloque
# ──────────────────────────────────────────────────────────────────────────────

ZONAS_VALIDAS = ('PICKING', 'RESERVA', 'AVERIAS', 'IMPORTADOS')
_PREFIJO_ZONA = {'PICKING': 'PIK', 'RESERVA': 'RES', 'AVERIAS': 'AVE', 'IMPORTADOS': 'IMP'}
_ZONAS_CUERPO = ('PICKING', 'RESERVA', 'IMPORTADOS')  # zonas que se arman con Mecanismo A (Cuerpo completo)
# Zonas donde el Hueco nunca se comparte: 1 SKU <-> 1 ubicación, pool de
# exclusividad independiente por zona (un SKU puede tener a la vez un slot en
# PICKING y otro en IMPORTADOS — no compiten entre sí, cada zona es su propia
# regla 1:1). RESERVA/AVERIAS quedan fuera: ahí N:N es libre.
_ZONAS_SLOT_UNICO = ('PICKING', 'IMPORTADOS')


def crear_cuerpo(almacen_id: int, pasillo: str, fila: int, cuerpo: int,
                 cantidad_entrepanos: int, tipo_zona: str, huecos_por_nivel: list = None):
    """
    Crea un Cuerpo completo de una vez: sus N Entrepaños (Nivel 1 = piso, subiendo)
    y, dentro de cada Entrepaño, sus Huecos. huecos_por_nivel es una lista de N
    enteros (uno por Entrepaño, en orden de Nivel) — cada Entrepaño declara su
    propia cantidad de Huecos porque la profundidad física varía entrepaño a
    entrepaño. Si no se pasa, cada Entrepaño nace con 1 Hueco (el caso normal:
    un SKU por Hueco).

    Este mecanismo solo arma el mapa físico vacío — el SKU y la capacidad de
    cada Hueco se amarran después, uno por uno, con asignar_producto().

    Un Cuerpo es 100% de una sola Zona (tipo_zona) — todos sus Entrepaños la
    heredan. PICKING, RESERVA e IMPORTADOS se arman como Cuerpos separados
    (Crear picking / Crear reserva / Crear importados en el front), no
    mezclados por Nivel dentro del mismo Cuerpo: así cada Cuerpo se revisa
    completo en una sola pestaña de zona.

    Dirección física: Pasillo -> Fila (1/2, lado del pasillo) -> Cuerpo (bahía)
    -> Nivel (entrepaño) -> Hueco.
    Código generado: {PREFIJO}-{PASILLO}{FILA}-C{CUERPO:02d}-E{NIVEL:02d}-H{HUECO:02d}
    ej. PIK-A1-C03-E02-H01 — la letra por eje (C/E/H = Cuerpo/Entrepaño/Hueco)
    evita confundir Cuerpo con Nivel al leer el código (antes eran dos dígitos
    seguidos, indistinguibles sin memorizar la posición).
    """
    if fila not in (1, 2):
        raise ValueError('fila debe ser 1 o 2 — cada pasillo tiene exactamente 2 filas')
    if cuerpo < 1:
        raise ValueError('cuerpo debe ser mayor a 0')
    if cantidad_entrepanos < 1:
        raise ValueError('cantidad_entrepanos debe ser mayor a 0')
    if tipo_zona not in _ZONAS_CUERPO:
        raise ValueError(f'tipo_zona debe ser una de {_ZONAS_CUERPO}')

    if huecos_por_nivel is None:
        huecos_por_nivel = [1] * cantidad_entrepanos
    if len(huecos_por_nivel) != cantidad_entrepanos:
        raise ValueError('huecos_por_nivel debe traer un valor por cada entrepaño')
    if any(h < 1 for h in huecos_por_nivel):
        raise ValueError('huecos_por_nivel: cada entrepaño necesita al menos 1 hueco')

    pasillo = pasillo.strip().upper()
    if not re.fullmatch(r'[A-Z]{1,2}', pasillo):
        raise ValueError('pasillo debe ser una letra o combinación A-Z / AA-ZZ')

    prefijo = _PREFIJO_ZONA[tipo_zona]
    creadas = []
    for nivel in range(1, cantidad_entrepanos + 1):
        for hueco in range(1, huecos_por_nivel[nivel - 1] + 1):
            codigo = f'{prefijo}-{pasillo}{fila}-C{cuerpo:02d}-E{nivel:02d}-H{hueco:02d}'
            if Ubicacion.query.filter_by(codigo=codigo).first():
                raise ValueError(f'La ubicación {codigo} ya existe')

            ub = Ubicacion(
                codigo=codigo,
                almacen_id=almacen_id,
                pasillo=pasillo,
                fila=fila,
                cuerpo=cuerpo,
                nivel=nivel,
                hueco=hueco,
                tipo_zona=tipo_zona,
                tipo='estanteria',
                origen='MANUAL',
                activo=True,
            )
            db.session.add(ub)
            creadas.append(ub)

    db.session.commit()
    return creadas


def editar_fila(almacen_id: int, pasillo: str, fila: int, tipo_zona: str = None,
                capacidad_maxima: int = None, activo: bool = None):
    """
    Edita en bloque todas las posiciones de una fila ya creada (mismo pasillo+fila).
    Reutiliza reclasificar_ubicacion() posición por posición para heredar sus
    guardarraíles (bloquea si hay stock activo, no aborta el lote si una posición
    falla — igual que importar_excel).
    """
    pasillo = pasillo.strip().upper()
    ubicaciones = Ubicacion.query.filter_by(
        almacen_id=almacen_id, pasillo=pasillo, estante=str(fila)
    ).order_by(Ubicacion.codigo).all()

    if not ubicaciones:
        raise ValueError(f'No hay posiciones creadas en {pasillo}{fila:02d}')

    actualizadas = []
    bloqueadas = {}
    advertencias = []
    for ub in ubicaciones:
        try:
            resultado = reclasificar_ubicacion(
                ubicacion_id=ub.id,
                tipo_zona=tipo_zona,
                capacidad_maxima=capacidad_maxima,
                activo=activo,
            )
            actualizadas.append(ub.codigo)
            advertencias.extend(resultado['advertencias'])
        except ValueError as e:
            bloqueadas[ub.codigo] = str(e)

    return {
        'pasillo': pasillo,
        'fila': fila,
        'total_posiciones': len(ubicaciones),
        'actualizadas': actualizadas,
        'bloqueadas': bloqueadas,
        'advertencias': advertencias,
    }


def _ubicaciones_de_cuerpo(almacen_id: int, pasillo: str, fila: int, cuerpo: int) -> list:
    """Todas las Ubicacion (Entrepaños/Huecos) de un Cuerpo — base de las 3 operaciones a nivel de Cuerpo."""
    pasillo = pasillo.strip().upper()
    ubicaciones = Ubicacion.query.filter_by(
        almacen_id=almacen_id, pasillo=pasillo, fila=fila, cuerpo=cuerpo,
    ).order_by(Ubicacion.nivel, Ubicacion.hueco).all()
    if not ubicaciones:
        raise ValueError(f'No existe el cuerpo {pasillo}{fila}-C{cuerpo:02d} en este almacén')
    return ubicaciones


# Tipos de MovimientoInventario que son puro bookkeeping interno del módulo
# Layout (asignar SKU, remodular cuerpo) — nunca representan que un operario
# haya pickeado o movido algo físicamente. editar_cuerpo() los ignora al
# decidir si puede remodular; eliminar_ubicacion()/eliminar_cuerpo() NO los
# ignoran (criterio más estricto y ya probado — ver _motivo_historial_bloqueante).
_TIPOS_MOVIMIENTO_BOOKKEEPING_LAYOUT = ('ASIGNACION_LAYOUT', 'REMODULACION_CUERPO')


def _motivo_referencias_nullable(ubicacion_id: int) -> str | None:
    """
    Tablas con FK nullable hacia ubicaciones que NO están en
    _motivo_historial_bloqueante/_motivo_historial_operativo_real por
    separado: ItemRecepcion (dos columnas: ubicacion_id y
    ubicacion_cross_dock_id), TareaDevolucion, LPN y LineaDevolucionCliente.
    Aunque nullable=True permite que el REGISTRO exista sin ubicación,
    Postgres igual bloquea el DELETE de la ubicación mientras una fila la
    referencie (regla NO ACTION) — así que también hay que revisarlas aquí,
    no solo las NOT NULL. Un LPN que sigue apuntando aquí es además una señal
    real: representa una estiba/pallet físico que puede seguir presente en
    el hueco.

    Descubierto en producción (2026-07-27): eliminar_cuerpo(forzar=True) sobre
    PIK-A1-C01 crasheaba con IntegrityError porque items_recepcion.ubicacion_id
    apuntaba ahí y nadie lo revisaba ni lo desvinculaba. Se verificó la lista
    completa de FKs reales contra information_schema.table_constraints en vez
    de confiar en memoria — salieron 8 tablas, no las 4 que este módulo conocía.

    Repetido en producción (2026-08-10), mismo síntoma y misma ubicación
    (PIK-A1-C01): lineas_devolucion_cliente.ubicacion_id (migración
    3ecf0c939454, módulo Devolución de Cliente) se creó DESPUÉS de esa
    auditoría y nunca se agregó aquí — 9na FK real, no 8. Recordatorio de la
    Regla 0 (una política, una función): cada FK nueva hacia ubicaciones debe
    sumarse a esta lista en el mismo commit que la crea, no esperar a que
    truene en producción para descubrirla.
    """
    if ItemRecepcion.query.filter(
        db.or_(
            ItemRecepcion.ubicacion_id == ubicacion_id,
            ItemRecepcion.ubicacion_cross_dock_id == ubicacion_id,
        )
    ).first():
        return 'tiene ítems de recepción asociados'
    if TareaDevolucion.query.filter_by(ubicacion_id=ubicacion_id).first():
        return 'tiene tareas de devolución asociadas'
    if LPN.query.filter_by(ubicacion_id=ubicacion_id).first():
        return 'tiene LPN(s) asociados'
    if LineaDevolucionCliente.query.filter_by(ubicacion_id=ubicacion_id).first():
        return 'tiene línea(s) de devolución de cliente asociada(s)'
    return None


def _motivo_historial_bloqueante(ubicacion_id: int) -> str | None:
    """
    Historial operativo que impide reconstruir o eliminar una ubicación sin
    perderlo — TareaPicking/TareaReposicion/SesionConteo tienen FK NOT NULL
    hacia ubicaciones. Usado por _motivo_no_eliminable() (que además bloquea
    por stock activo) para eliminar_ubicacion()/eliminar_cuerpo() — aquí
    CUALQUIER MovimientoInventario bloquea, incluida la sola asignación de un
    SKU, a propósito: eliminar borra la fila para siempre, así que el
    criterio es el más conservador posible (ver test_eliminar_fila_bloquea_
    por_historial_aunque_stock_sea_cero, que valida exactamente este caso).

    editar_cuerpo() NO usa esta función — usa _motivo_historial_operativo_real(),
    más permisiva, porque remodular no es tan destructivo como eliminar.
    """
    if TareaPicking.query.filter_by(ubicacion_id=ubicacion_id).first():
        return 'tiene historial de tareas de Picking'
    if TareaReposicion.query.filter(
        db.or_(
            TareaReposicion.ubicacion_picking_id == ubicacion_id,
            TareaReposicion.ubicacion_reserva_id == ubicacion_id,
        )
    ).first():
        return 'tiene historial de tareas de Reposición'
    if SesionConteo.query.filter_by(ubicacion_id=ubicacion_id).first():
        return 'tiene sesión de conteo cíclico asociada'
    if MovimientoInventario.query.filter_by(ubicacion_id=ubicacion_id).first():
        return 'tiene movimientos de inventario registrados'
    motivo_nullable = _motivo_referencias_nullable(ubicacion_id)
    if motivo_nullable:
        return motivo_nullable
    return None


def _motivo_historial_operativo_real(ubicacion_id: int) -> str | None:
    """
    Igual que _motivo_historial_bloqueante(), pero ignora los movimientos que
    son puro bookkeeping del propio módulo Layout (ver
    _TIPOS_MOVIMIENTO_BOOKKEEPING_LAYOUT). Usada por editar_cuerpo(): un hueco
    al que solo se le asignó un SKU (sin que ningún operario lo haya pickeado
    de verdad) sí se puede remodular — es exactamente el caso de uso real de
    "Editar": corregir la modulación de un cuerpo después de haber empezado a
    asignar SKUs, no solo antes de tocarlo.
    """
    if TareaPicking.query.filter_by(ubicacion_id=ubicacion_id).first():
        return 'tiene historial de tareas de Picking'
    if TareaReposicion.query.filter(
        db.or_(
            TareaReposicion.ubicacion_picking_id == ubicacion_id,
            TareaReposicion.ubicacion_reserva_id == ubicacion_id,
        )
    ).first():
        return 'tiene historial de tareas de Reposición'
    if SesionConteo.query.filter_by(ubicacion_id=ubicacion_id).first():
        return 'tiene sesión de conteo cíclico asociada'
    if MovimientoInventario.query.filter(
        MovimientoInventario.ubicacion_id == ubicacion_id,
        MovimientoInventario.tipo.notin_(_TIPOS_MOVIMIENTO_BOOKKEEPING_LAYOUT),
    ).first():
        return 'tiene movimientos de inventario registrados'
    motivo_nullable = _motivo_referencias_nullable(ubicacion_id)
    if motivo_nullable:
        return motivo_nullable
    return None


def _motivo_no_eliminable(ubicacion_id: int) -> str | None:
    """
    Ninguna posición con historial operativo se puede borrar — solo filas creadas
    por error y nunca usadas. Distinto (más estricto) que reclasificar_ubicacion:
    aquí no basta con stock=0 — hay 8 tablas con FK hacia ubicaciones (ver
    _motivo_historial_bloqueante y _motivo_referencias_nullable), algunas
    NOT NULL y otras nullable pero igual bloqueantes por la regla NO ACTION
    de Postgres — dejar huérfanos rompería el historial real de operación.
    """
    if _stock_activo(ubicacion_id) > 0:
        return 'tiene stock activo — muévelo antes de eliminar'
    motivo_historial = _motivo_historial_bloqueante(ubicacion_id)
    if motivo_historial:
        return f'{motivo_historial} — no se puede eliminar'
    return None


def _borrar_historial_de(ubicacion_id: int):
    """
    SOLO para el path forzar=True de eliminar_ubicacion()/eliminar_fila()/
    eliminar_cuerpo() — cuando el usuario pidió explícitamente saltarse el
    guardarraíl de historial (uso: pruebas, nunca el camino normal).

    Lista de las 9 tablas reales con FK hacia ubicaciones, verificada contra
    information_schema.table_constraints (no contra memoria — la primera
    versión de esta función solo conocía 4 y crasheó en producción dos veces
    seguidas con IntegrityError: primero por SesionConteo, después por
    ItemRecepcion, ambas con FK NOT NULL o NO ACTION que ninguna vía anterior
    revisaba; una tercera vez el 2026-08-10 por LineaDevolucionCliente, tabla
    creada después de esa auditoría). Dos estrategias según lo que permite
    cada esquema:

      - FK NOT NULL (no se puede desvincular, hay que borrar la fila entera):
        TareaPicking, TareaReposicion, SesionConteo.
      - FK nullable (el registro tiene sentido sin ubicación — se desvincula
        con UPDATE ... SET ubicacion_id = NULL, preservando el historial real
        en vez de borrarlo): MovimientoInventario, ItemRecepcion (dos
        columnas), LPN, TareaDevolucion, LineaDevolucionCliente.

    UbicacionProducto no aparece aquí — su FK también es NOT NULL, pero ya se
    borra aparte en cada llamador (eliminar_ubicacion/eliminar_fila/
    eliminar_cuerpo), antes o después de esta función.
    """
    TareaPicking.query.filter_by(ubicacion_id=ubicacion_id).delete(synchronize_session=False)
    TareaReposicion.query.filter(
        db.or_(
            TareaReposicion.ubicacion_picking_id == ubicacion_id,
            TareaReposicion.ubicacion_reserva_id == ubicacion_id,
        )
    ).delete(synchronize_session=False)
    SesionConteo.query.filter_by(ubicacion_id=ubicacion_id).delete(synchronize_session=False)

    MovimientoInventario.query.filter_by(ubicacion_id=ubicacion_id).update(
        {'ubicacion_id': None}, synchronize_session=False)
    ItemRecepcion.query.filter_by(ubicacion_id=ubicacion_id).update(
        {'ubicacion_id': None}, synchronize_session=False)
    ItemRecepcion.query.filter_by(ubicacion_cross_dock_id=ubicacion_id).update(
        {'ubicacion_cross_dock_id': None}, synchronize_session=False)
    LPN.query.filter_by(ubicacion_id=ubicacion_id).update(
        {'ubicacion_id': None}, synchronize_session=False)
    TareaDevolucion.query.filter_by(ubicacion_id=ubicacion_id).update(
        {'ubicacion_id': None}, synchronize_session=False)
    LineaDevolucionCliente.query.filter_by(ubicacion_id=ubicacion_id).update(
        {'ubicacion_id': None}, synchronize_session=False)


def eliminar_ubicacion(ubicacion_id: int, forzar: bool = False):
    """
    Elimina una sola ubicación que nunca se usó — versión de eliminar_fila()
    para una posición individual, con el mismo guardarraíl duro (bloquea si
    tiene stock o historial, no solo advierte).

    forzar=True se salta el guardarraíl completo (stock + historial) y borra
    en cascada TareaPicking/TareaReposicion/MovimientoInventario asociados —
    pérdida de datos permanente, pensado solo para limpiar pruebas, nunca para
    el flujo normal de operación.
    """
    ubicacion = Ubicacion.query.get(ubicacion_id)
    if not ubicacion:
        raise ValueError(f'Ubicación {ubicacion_id} no encontrada')

    if not forzar:
        motivo = _motivo_no_eliminable(ubicacion_id)
        if motivo:
            raise ValueError(f'{ubicacion.codigo} {motivo}')

    codigo = ubicacion.codigo
    if forzar:
        _borrar_historial_de(ubicacion_id)
    UbicacionProducto.query.filter_by(ubicacion_id=ubicacion.id).delete()
    db.session.delete(ubicacion)
    db.session.commit()
    return {'codigo': codigo}


def eliminar_cuerpo(almacen_id: int, pasillo: str, fila: int, cuerpo: int, forzar: bool = False) -> dict:
    """
    Elimina un Cuerpo completo — todos sus Entrepaños y Huecos, sin excepción.

    Todo o nada: si CUALQUIER hueco del cuerpo tiene stock activo o historial
    operativo (TareaPicking/TareaReposicion/MovimientoInventario), se bloquea
    la eliminación completa — no se borra parcialmente el cuerpo. Reutiliza
    _motivo_no_eliminable() por hueco (mismo guardarraíl duro que
    eliminar_ubicacion(), aplicado a los N huecos del cuerpo a la vez).

    Para retirar de operación un cuerpo con historial real sin perder su
    trazabilidad de auditoría, usar reclasificar_cuerpo(activo=False) en vez
    de esta función — desactivar archiva, eliminar borra para siempre.

    forzar=True se salta el guardarraíl completo en TODOS los huecos del
    cuerpo y borra en cascada su historial (ver _borrar_historial_de) —
    pérdida de datos permanente, pensado solo para limpiar pruebas.
    """
    ubicaciones = _ubicaciones_de_cuerpo(almacen_id, pasillo, fila, cuerpo)

    if not forzar:
        bloqueados = {}
        for ub in ubicaciones:
            motivo = _motivo_no_eliminable(ub.id)
            if motivo:
                bloqueados[ub.codigo] = motivo
        if bloqueados:
            detalle = '; '.join(f'{codigo} {motivo}' for codigo, motivo in bloqueados.items())
            raise ValueError(
                f'No se puede eliminar el cuerpo — {len(bloqueados)} de {len(ubicaciones)} '
                f'hueco(s) lo bloquean: {detalle}. Usa "Reclasificar > Desactivar cuerpo" '
                f'si necesitas retirarlo sin perder el historial.'
            )

    codigos = [ub.codigo for ub in ubicaciones]
    for ub in ubicaciones:
        if forzar:
            _borrar_historial_de(ub.id)
        UbicacionProducto.query.filter_by(ubicacion_id=ub.id).delete()
        db.session.delete(ub)
    db.session.commit()
    return {
        'pasillo': pasillo.strip().upper(), 'fila': fila, 'cuerpo': cuerpo,
        'eliminadas': codigos, 'total': len(codigos),
    }


def reclasificar_cuerpo(almacen_id: int, pasillo: str, fila: int, cuerpo: int,
                        tipo_zona: str = None, activo: bool = None) -> dict:
    """
    Reclasifica todo un Cuerpo de una vez: cambia su zona (RESERVA<->PICKING<->
    AVERIAS<->IMPORTADOS) o lo desactiva completo. Reutiliza reclasificar_ubicacion() hueco
    por hueco heredando su guardarraíl (bloquea por hueco si tiene stock
    activo), sin abortar el resto del cuerpo si una posición puntual falla —
    mismo patrón de "no aborta el lote" que editar_fila() ya usa para el
    esquema legado.

    activo=False es el camino recomendado para retirar de operación un cuerpo
    que eliminar_cuerpo() bloquea por tener historial real: no borra nada,
    solo lo saca de circulación — picking/conteo/traslados dejan de verlo
    porque ya no aparece activo — preservando la trazabilidad de auditoría
    para siempre.
    """
    ubicaciones = _ubicaciones_de_cuerpo(almacen_id, pasillo, fila, cuerpo)

    actualizadas = []
    bloqueadas = {}
    advertencias = []
    for ub in ubicaciones:
        try:
            resultado = reclasificar_ubicacion(
                ubicacion_id=ub.id, tipo_zona=tipo_zona, activo=activo,
            )
            actualizadas.append(ub.codigo)
            advertencias.extend(resultado['advertencias'])
        except ValueError as e:
            bloqueadas[ub.codigo] = str(e)

    return {
        'pasillo': pasillo.strip().upper(), 'fila': fila, 'cuerpo': cuerpo,
        'total_posiciones': len(ubicaciones),
        'actualizadas': actualizadas,
        'bloqueadas': bloqueadas,
        'advertencias': advertencias,
    }


def eliminar_fila(almacen_id: int, pasillo: str, fila: int, forzar: bool = False):
    """
    Elimina en bloque las posiciones de una fila que nunca se usaron. Pensado
    para deshacer una fila creada por error, no para dar de baja infraestructura
    en operación — por eso el guardarraíl es duro (bloquea, no solo advierte) y
    no aborta el lote si una posición sí tiene historial.

    forzar=True se salta el guardarraíl por posición y borra en cascada su
    historial (ver _borrar_historial_de) — pérdida de datos permanente,
    pensado solo para limpiar pruebas.
    """
    pasillo = pasillo.strip().upper()
    ubicaciones = Ubicacion.query.filter_by(
        almacen_id=almacen_id, pasillo=pasillo, estante=str(fila)
    ).order_by(Ubicacion.codigo).all()

    if not ubicaciones:
        raise ValueError(f'No hay posiciones creadas en {pasillo}{fila:02d}')

    eliminadas = []
    bloqueadas = {}
    for ub in ubicaciones:
        if not forzar:
            motivo = _motivo_no_eliminable(ub.id)
            if motivo:
                bloqueadas[ub.codigo] = motivo
                continue
        if forzar:
            _borrar_historial_de(ub.id)
        # Filas nunca usadas pueden tener registros de UbicacionProducto en 0
        # (ej. se asignó y luego se vació) — se limpian junto con la ubicación.
        UbicacionProducto.query.filter_by(ubicacion_id=ub.id).delete()
        eliminadas.append(ub.codigo)
        db.session.delete(ub)

    db.session.commit()
    return {
        'pasillo': pasillo,
        'fila': fila,
        'total_posiciones': len(ubicaciones),
        'eliminadas': eliminadas,
        'bloqueadas': bloqueadas,
    }


# ──────────────────────────────────────────────────────────────────────────────
# 3. AVERIAS — numeradas (AVE1, AVE2...), fuera de la grilla pasillo/fila
# ──────────────────────────────────────────────────────────────────────────────

def crear_ubicacion_averias(almacen_id: int, capacidad_maxima: int = None):
    """Crea la siguiente ubicación AVERIAS disponible (AVE1, AVE2...)."""
    existentes = Ubicacion.query.filter_by(
        almacen_id=almacen_id, tipo_zona='AVERIAS'
    ).all()

    numeros = []
    for ub in existentes:
        m = re.match(r'^AVE(\d+)$', ub.codigo, re.IGNORECASE)
        if m:
            numeros.append(int(m.group(1)))
    siguiente = (max(numeros) + 1) if numeros else 1
    codigo = f'AVE{siguiente}'

    ub = Ubicacion(
        codigo=codigo,
        almacen_id=almacen_id,
        tipo_zona='AVERIAS',
        tipo='estanteria',
        capacidad_maxima=capacidad_maxima,
        origen='MANUAL',
        activo=True,
    )
    db.session.add(ub)
    db.session.commit()
    return ub


# ──────────────────────────────────────────────────────────────────────────────
# 4. Mecanismo B — asignar(ubicacion_id, producto_id, cantidad)
# ──────────────────────────────────────────────────────────────────────────────

def _traspasar_desde_general(ubicacion: Ubicacion, producto_id: int, cantidad: int) -> int:
    """
    Fusión Layout↔Picking (solo NB1/CO003): al asignar un SKU a un hueco real,
    resta esa misma cantidad de SIESA-GENERAL (el bucket sin ubicación física
    donde aterriza el stock del sync de Siesa) para que picking, conteo cíclico
    y traslados —que ya leen UbicacionProducto sin filtrar por tipo_zona—
    empiecen a apuntar al hueco real en vez del bucket genérico.

    Las cantidades por ubicación son informativas (alertan reposición), no la
    fuente oficial de inventario — esa vive en Siesa. Por eso este traspaso
    nunca bloquea: resta como máximo lo que haya disponible en SIESA-GENERAL
    (piso en 0), y sigue de largo si no hay nada que mover.

    Retorna las unidades efectivamente movidas (0 si no aplicó).
    """
    if ubicacion.codigo == Ubicacion.CODIGO_GENERAL:
        return 0  # evita auto-traspaso si el hueco destino es el propio bucket

    almacen = Almacen.query.get(ubicacion.almacen_id)
    if not almacen or not almacen.tiene_fusion_layout_activa:
        return 0

    general = Ubicacion.query.filter_by(
        codigo=Ubicacion.CODIGO_GENERAL, almacen_id=ubicacion.almacen_id
    ).first()
    if not general:
        return 0

    reg_general = UbicacionProducto.query.filter_by(
        ubicacion_id=general.id, producto_id=producto_id, lote=None,
    ).with_for_update().first()
    if not reg_general or reg_general.cantidad <= 0:
        return 0

    movido = min(cantidad, reg_general.cantidad)
    reg_general.cantidad -= movido
    reg_general.row_version += 1
    return movido


def _traspasar_hacia_general(ubicacion: Ubicacion, producto_id: int, cantidad: int) -> int:
    """
    Inverso de _traspasar_desde_general() — usado por editar_cuerpo() antes de
    borrar un hueco al remodular: su stock físico real se devuelve a
    SIESA-GENERAL para que ninguna unidad se pierda en el proceso.

    A diferencia del traspaso de ida, esta salvaguarda es universal (no está
    restringida a NB1/CO003) y crea SIESA-GENERAL en el almacén si todavía no
    existe ahí — nunca puede fallar por falta de destino.

    Retorna las unidades movidas (siempre == cantidad, salvo cantidad <= 0).
    """
    if cantidad <= 0:
        return 0

    general = Ubicacion.query.filter_by(
        codigo=Ubicacion.CODIGO_GENERAL, almacen_id=ubicacion.almacen_id
    ).first()
    if not general:
        general = Ubicacion(
            codigo=Ubicacion.CODIGO_GENERAL, almacen_id=ubicacion.almacen_id,
            zona='GENERAL', tipo_zona='GENERAL', tipo='estanteria', activo=True,
        )
        db.session.add(general)
        db.session.flush()

    reg_general = UbicacionProducto.query.filter_by(
        ubicacion_id=general.id, producto_id=producto_id, lote=None,
    ).with_for_update().first()
    if reg_general:
        reg_general.cantidad += cantidad
        reg_general.row_version += 1
    else:
        db.session.add(UbicacionProducto(
            ubicacion_id=general.id, producto_id=producto_id, cantidad=cantidad,
            fecha_ingreso=datetime.utcnow(),
        ))
    return cantidad


def editar_cuerpo(almacen_id: int, pasillo: str, fila: int, cuerpo: int,
                  cantidad_entrepanos: int, huecos_por_nivel: list = None,
                  usuario_id: int = None) -> list:
    """
    Remodula un Cuerpo existente: cambia su cantidad de Entrepaños/Huecos.
    Borra todos los huecos actuales del cuerpo (y sus SKUs asignados) y los
    reconstruye desde cero con la nueva numeración — queda limpio y libre
    para volver a asignar SKUs, igual que un cuerpo recién creado. Conserva
    la zona original del cuerpo (editar remodula estructura, no cambia zona
    — para eso está reclasificar_cuerpo()).

    Antes de borrar cada hueco, su stock físico real se devuelve
    automáticamente a SIESA-GENERAL (_traspasar_hacia_general) — ninguna
    unidad se pierde en el proceso. Cada traspaso queda registrado en
    MovimientoInventario para trazabilidad.

    Guardarraíl duro: si CUALQUIER hueco del cuerpo tiene historial operativo
    REAL (TareaPicking, TareaReposicion, o un MovimientoInventario que no sea
    simple bookkeeping de Layout — ver _motivo_historial_operativo_real), se
    bloquea toda la remodulación. Haber asignado un SKU (ASIGNACION_LAYOUT)
    NO cuenta como historial bloqueante — es justamente el caso de uso normal
    de "Editar": corregir la modulación después de empezar a asignar SKUs, no
    solo antes. El stock tampoco bloquea (se traspasa); solo el historial de
    operación física real bloquea.
    """
    ubicaciones = _ubicaciones_de_cuerpo(almacen_id, pasillo, fila, cuerpo)
    tipo_zona = ubicaciones[0].tipo_zona  # un cuerpo es 100% de una sola zona

    if cantidad_entrepanos < 1:
        raise ValueError('cantidad_entrepanos debe ser mayor a 0')
    if huecos_por_nivel is None:
        huecos_por_nivel = [1] * cantidad_entrepanos
    if len(huecos_por_nivel) != cantidad_entrepanos:
        raise ValueError('huecos_por_nivel debe traer un valor por cada entrepaño')
    if any(h < 1 for h in huecos_por_nivel):
        raise ValueError('huecos_por_nivel: cada entrepaño necesita al menos 1 hueco')

    # Guardarraíl: bloquear TODO si algún hueco tiene historial operativo real.
    # Se valida antes de mutar nada — ningún traspaso ni borrado ocurre si falla.
    bloqueados = []
    for ub in ubicaciones:
        motivo = _motivo_historial_operativo_real(ub.id)
        if motivo:
            bloqueados.append(f'{ub.codigo}: {motivo}')
    if bloqueados:
        detalle = '; '.join(bloqueados)
        raise ValueError(
            f'No se puede remodular — {len(bloqueados)} de {len(ubicaciones)} hueco(s) '
            f'con historial real: {detalle}. Usa "Reclasificar > Desactivar cuerpo" '
            f'si necesitas retirarlo sin perder el historial.'
        )

    # Devolver stock físico de cada hueco a SIESA-GENERAL antes de borrar
    for ub in ubicaciones:
        regs = UbicacionProducto.query.filter_by(ubicacion_id=ub.id).with_for_update().all()
        for reg in regs:
            if reg.cantidad > 0:
                movido = _traspasar_hacia_general(ub, reg.producto_id, reg.cantidad)
                db.session.add(MovimientoInventario(
                    producto_id=reg.producto_id,
                    ubicacion_id=ub.id,
                    almacen_id=almacen_id,
                    tipo='REMODULACION_CUERPO',
                    cantidad=-movido,
                    saldo_antes=reg.cantidad,
                    saldo_despues=0,
                    motivo=f'Layout: {ub.codigo} remodulado — {movido} uds devueltas a SIESA-GENERAL',
                    usuario_id=usuario_id,
                    siesa_sync='OMITIDO',
                ))
                reg.cantidad = 0
                reg.row_version += 1

    # Borrar huecos actuales (ya sin stock ni historial — limpio)
    for ub in ubicaciones:
        UbicacionProducto.query.filter_by(ubicacion_id=ub.id).delete()
        db.session.delete(ub)
    db.session.flush()

    # Reconstruir desde cero con la nueva numeración (misma zona) — commitea al final
    return crear_cuerpo(
        almacen_id=almacen_id, pasillo=pasillo, fila=fila, cuerpo=cuerpo,
        cantidad_entrepanos=cantidad_entrepanos, tipo_zona=tipo_zona,
        huecos_por_nivel=huecos_por_nivel,
    )


def asignar_producto(ubicacion_id: int, producto_id: int, cantidad: int, usuario_id: int = None,
                     capacidad_maxima: int = None, stock_minimo: int = None):
    """
    Amarra un SKU a una ubicación y suma la cantidad contada.

    En PICKING e IMPORTADOS aplica la regla 1 SKU ↔ 1 ubicación por almacén,
    cada zona con su propio pool de exclusividad (un SKU puede tener a la vez
    un slot en PICKING y otro distinto en IMPORTADOS, no compiten entre sí):
    rechaza si la ubicación ya tiene otro SKU asignado, o si el SKU ya tiene
    otra ubicación de esa misma zona asignada en el mismo almacén — hay que
    liberar el slot primero (ver reclasificar_ubicacion).
    En RESERVA/AVERIAS no hay restricción — N:N libre.

    capacidad_maxima (opcional) solo aplica en PICKING/IMPORTADOS: como ahí el
    Hueco nunca se comparte con otro SKU, "capacidad del Hueco" y "capacidad
    para este SKU" son la misma cosa. En RESERVA/AVERIAS un Hueco puede tener
    varios SKUs a la vez, así que un solo valor de capacidad por Hueco no
    representa nada — se rechaza para no pisar silenciosamente el dato.

    stock_minimo (opcional, mismo alcance que capacidad_maxima) es el gatillo
    que lee reposicion_service.verificar_stock_picking() — sin este dato el
    hueco nunca genera TareaReposicion sin importar qué tan vacío quede.
    Antes había que asignarlo acá y volver a Reposición → Configurar a
    escribirlo por separado; ahora se hace en el mismo paso, delegando en
    reposicion_service.configurar_umbral() — Layout no valida ni escribe
    estos campos directamente, solo le pasa el dato a quien es dueño de esa
    regla (si capacidad_maxima también viene, se ofrece como techo por
    defecto de stock_maximo cuando la ubicación no tiene uno propio; ver
    esa función para el porqué).
    """
    if cantidad <= 0:
        raise ValueError('cantidad debe ser mayor a 0')

    ubicacion = Ubicacion.query.get(ubicacion_id)
    if not ubicacion:
        raise ValueError(f'Ubicación {ubicacion_id} no encontrada')
    if not ubicacion.activo:
        raise ValueError(f'Ubicación {ubicacion.codigo} está inactiva')

    producto = Producto.query.get(producto_id)
    if not producto:
        raise ValueError(f'Producto {producto_id} no encontrado')

    if capacidad_maxima is not None and ubicacion.tipo_zona not in _ZONAS_SLOT_UNICO:
        raise ValueError(
            f'capacidad_maxima solo aplica a Huecos {"/".join(_ZONAS_SLOT_UNICO)} — en '
            f'RESERVA/AVERIAS el Hueco puede compartirse entre varios SKUs'
        )

    if ubicacion.tipo_zona in _ZONAS_SLOT_UNICO:
        if ubicacion.producto_asignado_id and ubicacion.producto_asignado_id != producto_id:
            otro = Producto.query.get(ubicacion.producto_asignado_id)
            raise ValueError(
                f'{ubicacion.codigo} ya está asignada a {otro.codigo if otro else ubicacion.producto_asignado_id} '
                f'— libera el slot antes de reasignarlo'
            )
        otra_ub = Ubicacion.query.filter(
            Ubicacion.almacen_id == ubicacion.almacen_id,
            Ubicacion.tipo_zona == ubicacion.tipo_zona,
            Ubicacion.producto_asignado_id == producto_id,
            Ubicacion.id != ubicacion.id,
        ).first()
        if otra_ub:
            raise ValueError(
                f'{producto.codigo} ya tiene un slot de {ubicacion.tipo_zona} asignado en {otra_ub.codigo} '
                f'— libéralo antes de asignar uno nuevo'
            )
        ubicacion.producto_asignado_id = producto_id

    if capacidad_maxima is not None:
        ubicacion.capacidad_maxima = capacidad_maxima  # dato propio de Layout: cuánto cabe físicamente

    if stock_minimo is not None or capacidad_maxima is not None:
        # Layout no conoce las reglas de Reposición (zona válida, mínimo ≤
        # máximo, cuándo usar capacidad_maxima como techo por defecto) —
        # se las delega a quien es dueño de ellas. Ver configurar_umbral().
        from app.services import reposicion_service as _reposicion
        umbral_kwargs = {}
        if stock_minimo is not None:
            umbral_kwargs['stock_minimo'] = stock_minimo
        if capacidad_maxima is not None:
            umbral_kwargs['capacidad_referencia'] = capacidad_maxima
        _reposicion.configurar_umbral(ubicacion.id, **umbral_kwargs)

    reg = UbicacionProducto.query.filter_by(
        ubicacion_id=ubicacion_id, producto_id=producto_id, lote=None,
    ).with_for_update().first()

    if not reg:
        reg = UbicacionProducto(
            ubicacion_id=ubicacion_id, producto_id=producto_id, cantidad=0,
            fecha_ingreso=datetime.utcnow(),
        )
        db.session.add(reg)
        db.session.flush()

    saldo_antes = reg.cantidad
    reg.cantidad += cantidad
    reg.row_version += 1

    # Lock de destino ya tomado arriba (reg) — ahora sí el de origen (SIESA-GENERAL),
    # mismo orden que confirmar_reposicion() para evitar deadlocks cruzados.
    movido_de_general = _traspasar_desde_general(ubicacion, producto_id, cantidad)

    motivo = f'Layout: {producto.codigo} asignado a {ubicacion.codigo}'
    if movido_de_general:
        motivo += f' (traspaso {movido_de_general} desde SIESA-GENERAL)'

    db.session.add(MovimientoInventario(
        producto_id=producto_id,
        ubicacion_id=ubicacion_id,
        almacen_id=ubicacion.almacen_id,
        tipo='ASIGNACION_LAYOUT',
        cantidad=cantidad,
        saldo_antes=saldo_antes,
        saldo_despues=reg.cantidad,
        motivo=motivo,
        usuario_id=usuario_id,
        siesa_sync='OMITIDO',  # movimiento 100% interno del WMS, Siesa no se entera
    ))

    db.session.commit()
    return {
        'ubicacion': ubicacion.to_dict(),
        'producto_id': producto_id,
        'producto_codigo': producto.codigo,
        'cantidad_total': reg.cantidad,
    }


# ──────────────────────────────────────────────────────────────────────────────
# 5. Reclasificación — con guardarraíles
# ──────────────────────────────────────────────────────────────────────────────

def _stock_activo(ubicacion_id: int) -> int:
    return db.session.query(
        db.func.coalesce(db.func.sum(UbicacionProducto.cantidad), 0)
    ).filter_by(ubicacion_id=ubicacion_id).scalar()


def reclasificar_ubicacion(ubicacion_id: int, tipo_zona: str = None,
                            capacidad_maxima: int = None, activo: bool = None,
                            liberar_slot: bool = False):
    """
    Cambia zona/capacidad/estado de una ubicación ya existente.

    Guardarraíles:
      - No reclasifica zona ni desactiva si hay stock activo (cantidad > 0) —
        hay que mover ese stock antes.
      - Avisa (no bloquea) si hay TareaPicking/TareaReposicion PENDIENTE/EN_PROCESO
        apuntando a esta ubicación.
    """
    ubicacion = Ubicacion.query.get(ubicacion_id)
    if not ubicacion:
        raise ValueError(f'Ubicación {ubicacion_id} no encontrada')

    stock = _stock_activo(ubicacion_id)
    cambia_zona = tipo_zona is not None and tipo_zona != ubicacion.tipo_zona
    desactiva = activo is False and ubicacion.activo

    if (cambia_zona or desactiva) and stock > 0:
        raise ValueError(
            f'{ubicacion.codigo} tiene {stock} unidades activas — muévelas antes de '
            f'reclasificar o desactivar esta ubicación'
        )

    if liberar_slot:
        if stock > 0:
            raise ValueError(
                f'{ubicacion.codigo} tiene {stock} unidades activas — no se puede '
                f'liberar el slot con stock presente'
            )
        ubicacion.producto_asignado_id = None

    if tipo_zona is not None:
        if tipo_zona not in ZONAS_VALIDAS and tipo_zona != 'GENERAL':
            raise ValueError(f'tipo_zona debe ser una de {ZONAS_VALIDAS} o GENERAL')
        ubicacion.tipo_zona = tipo_zona
    if capacidad_maxima is not None:
        ubicacion.capacidad_maxima = capacidad_maxima
    if activo is not None:
        ubicacion.activo = activo

    advertencias = []
    tareas_picking = TareaPicking.query.filter_by(
        ubicacion_id=ubicacion_id
    ).filter(TareaPicking.estado.in_(['PENDIENTE', 'EN_PROCESO'])).count()
    if tareas_picking:
        advertencias.append(f'{tareas_picking} tarea(s) de Picking pendiente(s) en esta ubicación')

    tareas_reposicion = TareaReposicion.query.filter(
        db.or_(
            TareaReposicion.ubicacion_picking_id == ubicacion_id,
            TareaReposicion.ubicacion_reserva_id == ubicacion_id,
        ),
        TareaReposicion.estado.in_(['PENDIENTE', 'EN_PROCESO']),
    ).count()
    if tareas_reposicion:
        advertencias.append(f'{tareas_reposicion} tarea(s) de Reposición pendiente(s) en esta ubicación')

    db.session.commit()
    return {'ubicacion': ubicacion.to_dict(), 'advertencias': advertencias}


# ──────────────────────────────────────────────────────────────────────────────
# 6. Mecanismo A (opción masiva) — importador Excel
# ──────────────────────────────────────────────────────────────────────────────

def importar_excel(almacen_id: int, file_stream, usuario_id: int = None):
    """
    Carga masiva: columnas ubicacion_codigo | producto_codigo | cantidad.
    No aborta en la primera fila mala — reporta éxito/error fila por fila,
    igual que necesita una migración de miles de SKUs desde SIESA-GENERAL.
    """
    from openpyxl import load_workbook

    wb = load_workbook(file_stream, read_only=True, data_only=True)
    ws = wb.active

    filas = list(ws.iter_rows(min_row=2, values_only=True))
    resultados = {'ok': 0, 'errores': []}

    for i, fila in enumerate(filas, start=2):
        if not fila or all(c is None for c in fila):
            continue
        try:
            codigo_ub, codigo_prod, cantidad = fila[0], fila[1], fila[2]
            if not codigo_ub or not codigo_prod or cantidad is None:
                raise ValueError('faltan columnas (ubicacion_codigo, producto_codigo, cantidad)')

            ubicacion = Ubicacion.query.filter_by(
                codigo=str(codigo_ub).strip(), almacen_id=almacen_id
            ).first()
            if not ubicacion:
                raise ValueError(f'ubicación "{codigo_ub}" no existe en este almacén')

            producto = Producto.query.filter_by(codigo=str(codigo_prod).strip()).first()
            if not producto:
                raise ValueError(f'producto "{codigo_prod}" no existe')

            asignar_producto(ubicacion.id, producto.id, int(cantidad), usuario_id)
            resultados['ok'] += 1
        except Exception as e:
            resultados['errores'].append({'fila': i, 'error': str(e)})

    return resultados
